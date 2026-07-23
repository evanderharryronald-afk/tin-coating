"""
MergePipeline: 高性能主副表合并工具

核心功能：
  - 解析复杂 Excel 表头（支持合并单元格、多级标题）
  - 本地缓存优化（避免重复解析）
  - 前置过滤 + 向量化聚合
  - 支持多副表级联合并
"""

import pandas as pd
import numpy as np
import yaml
import json
import hashlib
import time
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class MergePipeline:
    """高性能主副表合并管道"""

    def __init__(self, config_path: str, base_dir: str = None):
        """初始化合并管道"""
        self.config_path = Path(config_path)
        self.base_dir = Path(base_dir) if base_dir else self.config_path.parent
        
        if not self.config_path.exists():
            raise FileNotFoundError(f"Config file not found: {self.config_path}")
        
        with open(self.config_path, encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
        
        self.master_cfg = self.config['master']
        self.header_row_start = self.master_cfg.get('header_row_start', 11)
        self.header_row_end = self.master_cfg.get('header_row_end', 14)
        self.data_start_row = self.master_cfg.get('data_start_row', 15)
        
        print("[MergePipeline] Initialized")

    # ── Utility functions ──
    def _build_merge_map(self, ws):
        """Expand all merged cells: {(row, col): top-left value}"""
        m = {}
        for mr in ws.merged_cells.ranges:
            top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    m[(r, c)] = top_val
        return m

    def _get_val(self, ws, merge_map, row, col):
        """Get value from merge map or cell"""
        return merge_map.get((row, col), ws.cell(row=row, column=col).value)

    def _ffill(self, lst):
        """Forward fill empty values in list"""
        result, last = [], None
        for v in lst:
            sv = str(v).strip() if v is not None else ''
            if sv:
                last = sv
            result.append(last)
        return result

    def _clean_str(self, s):
        """Clean string"""
        if s is None:
            return ''
        return str(s).replace('\n', '').replace('\r', '').strip()

    def _detect_max_col(self, ws, merge_map):
        """Detect maximum column number"""
        max_col = max((c for r, c in merge_map.keys()
                       if self.header_row_start <= r <= self.header_row_end), default=0)
        for r in range(self.header_row_start, self.header_row_end + 1):
            for c in range(max_col, max_col + 20):
                v = ws.cell(row=r, column=c).value
                if v is not None and self._clean_str(v):
                    max_col = max(max_col, c)
        return max_col

    def _generate_column_names(self, ws, merge_map, max_col):
        """Generate final column names from multi-level headers"""
        raw = []
        for row_idx in range(self.header_row_start, self.header_row_end + 1):
            raw.append([self._get_val(ws, merge_map, row_idx, c) for c in range(1, max_col + 1)])

        l1f = self._ffill(raw[0])
        l2f = self._ffill(raw[1])
        l3f = self._ffill(raw[2])
        l4 = [self._clean_str(v) for v in raw[3]]

        col_names = []
        for i in range(max_col):
            parts = []
            for lvl in [l1f[i], l2f[i], l3f[i], l4[i]]:
                sv = self._clean_str(lvl)
                if sv and (not parts or parts[-1] != sv):
                    parts.append(sv)
            col_names.append('_'.join(parts) if parts else f'col_{i + 1:03d}')

        # Handle duplicate names
        counts = Counter()
        final = []
        for name in col_names:
            counts[name] += 1
            final.append(f"{name}_{counts[name] - 1}" if counts[name] > 1 else name)
        return final

    # ── Core methods ──
    def parse_master_schema(self, path: Path, keep_col_letters: list):
        """Parse master table structure and generate column names"""
        print(f"[PARSE] Analyzing master table: {path.parent.name}/{path.name}")
        
        wb_full = load_workbook(str(path), read_only=False, data_only=True)
        ws_full = wb_full.active
        merge_map = self._build_merge_map(ws_full)
        max_col = self._detect_max_col(ws_full, merge_map)
        print(f"[PARSE] Detected {max_col} columns")
        
        all_col_names = self._generate_column_names(ws_full, merge_map, max_col)
        wb_full.close()

        # Apply column name corrections if available
        patch_path = self.base_dir / "rename_patch.yaml"
        if patch_path.exists():
            try:
                with open(patch_path, encoding='utf-8') as f:
                    rename_patch = yaml.safe_load(f)
                if rename_patch and isinstance(rename_patch, dict):
                    for idx, name in enumerate(all_col_names):
                        if name in rename_patch:
                            correct_name = rename_patch[name]
                            print(f"[PATCH] Column rename: [{name}] -> [{correct_name}]")
                            all_col_names[idx] = correct_name
            except Exception as e:
                print(f"[WARN] Failed to load rename_patch.yaml: {e}")

        # Map keep columns to indices
        keep_indices = []
        col_letter_map = {}
        for letter in keep_col_letters:
            letter = letter.strip().upper()
            idx = column_index_from_string(letter) - 1
            if idx >= max_col:
                raise RuntimeError(f"Column {letter} exceeds max ({max_col})")
            semantic_name = all_col_names[idx]
            keep_indices.append(idx)
            col_letter_map[semantic_name] = letter

        selected_names = [all_col_names[i] for i in keep_indices]
        print(f"[PARSE] Schema parsed, keeping {len(keep_indices)} columns")
        return keep_indices, selected_names, col_letter_map

    def get_master_schema_with_cache(self, master_path: Path, keep_col_letters: list):
        """Get master schema with local file caching"""
        cache_path = self.base_dir / "master_schema_cache.json"
        patch_path = self.base_dir / "rename_patch.yaml"

        patch_content = ""
        if patch_path.exists():
            patch_content = patch_path.read_text(encoding='utf-8')

        config_str = f"{master_path.name}_{self.header_row_start}_{self.header_row_end}_{','.join(keep_col_letters)}_{patch_content}"
        config_hash = hashlib.md5(config_str.encode('utf-8')).hexdigest()

        # Try to load from cache
        if cache_path.exists():
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    cache_data = json.load(f)
                if cache_data.get("config_hash") == config_hash:
                    print("[CACHE] Found matching master schema cache")
                    return cache_data["keep_indices"], cache_data["selected_names"], cache_data["col_letter_map"]
            except Exception:
                print("[WARN] Failed to read cache, will re-parse master table")

        # Parse master schema
        keep_indices, selected_names, col_letter_map = self.parse_master_schema(master_path, keep_col_letters)

        # Save to cache
        try:
            cache_data = {
                "config_hash": config_hash,
                "keep_indices": keep_indices,
                "selected_names": selected_names,
                "col_letter_map": col_letter_map
            }
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
            print("[CACHE] Schema cached")
        except Exception as e:
            print(f"[WARN] Failed to cache: {e}")

        return keep_indices, selected_names, col_letter_map

    def load_master_data(self, path: Path, keep_indices: list, selected_names: list, col_letter_map: dict) -> pd.DataFrame:
        """Load master table data"""
        print(f"[LOAD] Reading master table: {path.parent.name}/{path.name}")
        
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(min_row=self.data_start_row, values_only=True):
            if all(v is None for v in row):
                break
            rows.append([row[i] if i < len(row) else None for i in keep_indices])

        wb.close()

        if not rows:
            raise RuntimeError(f"No data rows found in {path.name}")

        df = pd.DataFrame(rows, columns=selected_names)
        df.attrs['col_letter_map'] = col_letter_map
        print(f"[LOAD] Loaded {len(df)} rows x {len(df.columns)} columns")
        return df

    def _load_sub(self, path: Path) -> pd.DataFrame:
        """Load sub table (supports xlsx/xls/csv)"""
        suffix = path.suffix.lower()
        if suffix in ('.xlsx', '.xls'):
            return pd.read_excel(str(path), header=0)
        elif suffix == '.csv':
            try:
                return pd.read_csv(str(path), header=0, encoding='utf-8')
            except UnicodeDecodeError:
                return pd.read_csv(str(path), header=0, encoding='gbk')
        else:
            raise RuntimeError(f"Unsupported format: {path.suffix}")

    def merge_one_sub(self, master_df: pd.DataFrame, sub_cfg: dict) -> pd.DataFrame:
        """High-performance sub table merge with pre-filtering and vectorized aggregation"""
        sub_path_str = sub_cfg['path'].replace('\\', '/')
        sub_path = self.base_dir / sub_path_str
        
        if not sub_path.exists():
            raise RuntimeError(f"Sub table not found: {sub_path}")

        print(f"[MERGE] Reading sub table: {sub_path.name}")
        sub_df = self._load_sub(sub_path)
        print(f"[MERGE] Original: {len(sub_df)} rows x {len(sub_df.columns)} columns")

        # 1. Locate key fields
        join_key = sub_cfg['join_key']
        master_key_letter = join_key['master_col'].strip().upper()
        sub_key_col = join_key['sub_col'].strip()

        master_key_col = None
        for col in master_df.columns:
            if master_df.attrs.get('col_letter_map', {}).get(col) == master_key_letter:
                master_key_col = col
                break
        
        if master_key_col is None:
            raise RuntimeError(f"Cannot find column {master_key_letter} in master table")

        if sub_key_col not in sub_df.columns:
            raise RuntimeError(f"Cannot find column {sub_key_col} in sub table")

        # Convert to string for matching
        master_df[master_key_col] = master_df[master_key_col].astype(str).str.strip()
        sub_df[sub_key_col] = sub_df[sub_key_col].astype(str).str.strip()

        # 2. Pre-filter: remove null required columns
        required_cols_cfg = join_key.get('required_columns', [])
        if required_cols_cfg:
            valid_req_cols = [c for c in required_cols_cfg if c in sub_df.columns]
            if valid_req_cols:
                before = len(sub_df)
                sub_df = sub_df.dropna(subset=valid_req_cols, how='any')
                print(f"[OPT] Pre-filter: {before} -> {len(sub_df)} rows")

        # Remove rows not in master
        master_keys_set = set(master_df[master_key_col].unique())
        before = len(sub_df)
        sub_df = sub_df[sub_df[sub_key_col].isin(master_keys_set)]
        print(f"[OPT] Key filter: {before} -> {len(sub_df)} rows")

        if len(sub_df) == 0:
            print("[WARN] No valid rows after filtering")
            return master_df

        # 3. Build aggregation dict
        keep_cols_cfg = sub_cfg.get('keep_columns', [])
        if not keep_cols_cfg:
            print(f"[WARN] No keep_columns configured for {sub_path.name}")
            return master_df

        agg_dict = {}
        rename_dict = {}
        master_cols = set(master_df.columns)

        strategy_map = {
            'mean': 'mean', 'sum': 'sum', 'max': 'max', 'min': 'min',
            'first': 'first', 'first_nonempty': 'first'
        }

        for item in keep_cols_cfg:
            col_name = item['col'].strip()
            if col_name not in sub_df.columns:
                raise RuntimeError(f"Cannot find column {col_name} in sub table")

            strategy = item.get('strategy', 'mean')
            agg_op = strategy_map.get(strategy, 'mean')

            if strategy == 'join_unique':
                agg_dict[col_name] = lambda x: ','.join(x.dropna().astype(str).unique())
            else:
                agg_dict[col_name] = agg_op

            if col_name in master_cols and col_name != master_key_col:
                rename_dict[col_name] = f"{col_name}_sub"

        # 4. Vectorized aggregation
        print("[OPT] Vectorized aggregation...")
        agg_df = sub_df.groupby(sub_key_col, as_index=False).agg(agg_dict)

        if rename_dict:
            agg_df = agg_df.rename(columns=rename_dict)

        # 5. Merge
        unmatched_action = join_key.get('unmatched', 'keep').strip().lower()
        join_how = 'inner' if unmatched_action == 'drop' else 'left'

        result = master_df.merge(agg_df, left_on=master_key_col, right_on=sub_key_col, how=join_how)

        if sub_key_col != master_key_col and sub_key_col in result.columns:
            result = result.drop(columns=[sub_key_col])

        result.attrs['col_letter_map'] = master_df.attrs.get('col_letter_map', {})
        print(f"[MERGE] Merged: {len(result)} rows")
        return result

    # ── Main entry point ──
    def run(self) -> pd.DataFrame:
        """Execute full merge pipeline"""
        t_start = time.time()

        # 1. Locate master tables
        search_root_str = self.master_cfg['search_root'].replace('\\', '/')
        search_root = self.base_dir / search_root_str
        pattern = self.master_cfg.get('search_pattern', f"*/{self.master_cfg['filename']}")
        master_paths = sorted(search_root.glob(pattern))

        if not master_paths:
            raise RuntimeError(f"No master tables found in {search_root}")

        print(f"Found {len(master_paths)} master tables")

        keep_col_letters = self.master_cfg.get('keep_columns', [])
        if not keep_col_letters:
            raise RuntimeError("No keep_columns configured")

        # 2. Get schema with cache
        keep_indices, selected_names, col_letter_map = self.get_master_schema_with_cache(
            master_paths[0], keep_col_letters
        )

        # 3. Load and concatenate all master tables
        all_results = []
        for master_path in master_paths:
            result_df = self.load_master_data(master_path, keep_indices, selected_names, col_letter_map)
            all_results.append(result_df)

        print(f"Concatenating {len(all_results)} tables...")
        final_df = pd.concat(all_results, ignore_index=True)
        final_df.attrs['col_letter_map'] = col_letter_map

        # 4. Merge sub tables one by one
        sub_tables = self.config.get('sub_tables', [])
        for i, sub_cfg in enumerate(sub_tables, 1):
            print(f"\n-- Merging sub table {i} --")
            final_df = self.merge_one_sub(final_df, sub_cfg)

        # 5. Write output
        out_cfg = self.config.get('output', {})
        out_path_str = out_cfg.get('path', 'merged_result.xlsx').replace('\\', '/')
        out_path = self.base_dir / out_path_str
        out_sheet = out_cfg.get('sheet_name', 'Result')

        print(f"\n[SAVE] Writing to: {out_path.name}")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name=out_sheet, index=False)

        elapsed = time.time() - t_start
        print(f"\nDONE: {out_path}")
        print(f"  {len(final_df)} rows x {len(final_df.columns)} columns")
        print(f"  Time: {elapsed:.1f}s\n")

        return final_df
