"""
SurfaceEDAAnalyzer (generalized)
================================
清洗后 / 建模前 / 建模后 的分布与影响分析模块。

注意：本模块中的 "delta" 实际上指的是目标变量 Delta（偏差量 = 实验室值 - 在线测量值），
而不是模型预测残差。Delta > 0 表示在线测量偏低，Delta < 0 表示在线测量偏高。

设计目标（相对初版的增强）：
- 不原地修改传入的 DataFrame
- 列名与衍生特征可配置，硬编码降到最低
- 支持 delta_col / delta Series / (y_true + y_pred) 三种输入方式
- 统计与绘图解耦（可只算统计、不画图）
- 支持 target_groups 白名单，与主流程达标组对齐
- 返回统一 summary DataFrame，方便写入现有 Excel 报表
- 保持 surface 快捷方式，兼容旧调用
"""

from __future__ import annotations

import os
import warnings
from typing import Any, Callable, Dict, List, Optional, Sequence, Union
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from scipy import stats
from sympy.codegen.fnodes import dimension

# 中文字体兜底（环境若无对应字体则自动忽略）
plt.rcParams["axes.unicode_minus"] = False
try:
    plt.rcParams["font.sans-serif"] = ["SimHei", "Arial Unicode MS", "DejaVu Sans"]
except Exception:
    pass


class SurfaceEDAAnalyzer:
    """模块化：特征分布 + 残差分布 + 特征对残差影响 + 时间趋势 + 分规格 + train/test 漂移"""

    def __init__(self, default_save_dir: str = "result/eda_result"):
        self.default_save_dir = default_save_dir

    # ------------------------------------------------------------------
    # 公共入口
    # ------------------------------------------------------------------
    def analyze(
        self,
        df: pd.DataFrame,
        surface: Optional[str] = None,  # 'Top' / 'Bot'，可选，用于自动拼列名
        delta_col: Optional[str] = None,  # 直接指定残差列（优先）
        delta: Optional[pd.Series] = None,  # 或直接传 delta Series
        y_true: Optional[pd.Series] = None,  # 或传 y_true + y_pred 计算残差
        y_pred: Optional[pd.Series] = None,
        lab_col: Optional[str] = None,
        actual_col: Optional[str] = None,
        feature_cols: Optional[Sequence[str]] = None,
        feature_factory: Optional[Callable[[pd.DataFrame], pd.DataFrame]] = None,
        time_col: str = "Produce Time",
        group_col: Optional[str] = None,  # 分规格列，例如 Setpoint_Group_Label
        target_groups: Optional[Sequence[str]] = None,  # 白名单，仅分析这些组
        model_residual_col: Optional[str] = None,  # 兼容旧接口：建模后残差列
        train_idx: Optional[Union[pd.Index, np.ndarray, List]] = None,
        test_idx: Optional[Union[pd.Index, np.ndarray, List]] = None,
        save_dir: Optional[str] = None,
        max_groups: int = 12,
        sample_for_scatter: Optional[int] = 8000,
        plot_univariate: bool = True,
        plot_vs_delta: bool = True,
        plot_time: bool = True,
        plot_train_test: bool = True,
        plot_model_residual: bool = True,
        compute_stats_only: bool = False,  # True 时只算统计、不画图
        figsize_univariate: tuple = (8, 4),
        random_state: int = 42,
        # 镀层场景默认特征覆盖（可选）
        default_feature_map: Optional[Dict[str, List[str]]] = None,
        enable_overall: bool = True,      # 是否做整体分析
        enable_by_group: bool = True,     # 是否做分组分析
    ) -> Dict:
        """
        主分析入口。

        残差优先级：
        1. delta (Series)
        2. delta_col (列名)
        3. y_true - y_pred
        4. lab_col - actual_col（或由 surface 自动推导）

        Parameters
        ----------
        df : 输入 DataFrame（不会被原地修改）
        surface : 'Top' 或 'Bot'，传入后自动构造 lab/actual 等列
        delta / delta_col / y_true+y_pred / lab_col+actual_col : 残差来源
        feature_cols : 要分析的特征列表；None 时按 surface 或 default_feature_map 推导
        feature_factory : 可选回调，接收 copy 后的 df，返回带衍生特征的 df
        group_col / target_groups : 分规格与白名单
        compute_stats_only : 只计算统计、不画图（适合调参循环）
        default_feature_map : 例如 {"Top": [...], "Bot": [...]}，覆盖内置默认特征

        Returns
        -------
        dict : 包含 delta_col、feature_cols、n_samples、paths、stats、summary_df

        enable_overall : bool, default=True   是否执行整体（全量样本）EDA 分析
        enable_by_group : bool, default=True  是否执行分规格 EDA 分析（需要 group_col 不为 None）

        """
        out_dir = save_dir if save_dir is not None else self.default_save_dir
        if not compute_stats_only:
            os.makedirs(out_dir, exist_ok=True)

        # ---------- 1. 工作副本，禁止原地修改 ----------
        work = df.copy()

        # ---------- 2. 解析列名 + 衍生特征 ----------
        lab_col, actual_col, default_features, surface_cn = self._resolve_columns(
            work, surface, lab_col, actual_col, feature_cols, default_feature_map
        )

        # 用户自定义衍生特征工厂（在 copy 上操作）
        if feature_factory is not None:
            work = feature_factory(work)

        # 内置衍生：Current_Per_Speed（仅当 surface 可用且列存在时）
        if surface is not None:
            prefix = "Top" if surface == "Top" else "Bot"
            speed_col = "Speed[m/min]_Process_Avg"
            current_col = f"{prefix}_Current_Sum"
            per_speed = f"{prefix}_Current_Per_Speed"
            if current_col in work.columns and speed_col in work.columns and per_speed not in work.columns:
                work[per_speed] = work[current_col] / (work[speed_col] + 1e-5)

        # ---------- 3. 确定残差列 ----------
        delta_name = "_delta_"
        if delta is not None:
            delta = delta.reindex(work.index)
            work[delta_name] = delta
            delta_col = delta_name
            print(f"[测量偏差/Delta] 使用传入的 delta Series")
        elif delta_col is not None:
            if delta_col not in work.columns:
                raise ValueError(f"指定的 delta_col={delta_col} 不在 DataFrame 中")
            print(f"[测量偏差/Delta] 使用已有列: {delta_col}")
        elif y_true is not None and y_pred is not None:
            y_true = y_true.reindex(work.index)
            y_pred = y_pred.reindex(work.index)
            work[delta_name] = y_true - y_pred
            delta_col = delta_name
            print(f"[测量偏差/Delta] 使用 y_true - y_pred 计算残差")
        else:
            if lab_col is None or actual_col is None:
                raise ValueError(
                    "未提供 delta / delta_col / (y_true+y_pred)，"
                    "且无法从 lab_col / actual_col 计算残差。"
                    "请至少传入其中一种，或传入 surface。"
                )
            if lab_col not in work.columns or actual_col not in work.columns:
                raise ValueError(f"lab_col={lab_col} 或 actual_col={actual_col} 不在 DataFrame 中")
            work[delta_name] = work[lab_col] - work[actual_col]
            delta_col = delta_name
            print(f"[测量偏差/Delta] 使用 {lab_col} - {actual_col} 计算残差")

        # 特征列表最终确认
        if feature_cols is None:
            feature_cols = default_features
        feature_cols = [c for c in feature_cols if c in work.columns and c != delta_col]
        if not feature_cols:
            warnings.warn("没有可用的特征列，仅做残差自身分布分析")

        # 时间列处理
        has_time = time_col in work.columns
        if has_time:
            work[time_col] = pd.to_datetime(work[time_col], errors="coerce")

        # 统一有效样本（残差非空）
        base_cols = [delta_col] + list(feature_cols)
        if has_time:
            base_cols.append(time_col)
        if group_col and group_col in work.columns:
            base_cols.append(group_col)
        # 去重并保留顺序
        seen = set()
        base_cols = [c for c in base_cols if not (c in seen or seen.add(c))]
        data = work[base_cols].dropna(subset=[delta_col]).copy()
        print(f"[数据] 有效样本数（残差非空）: {len(data)}")

        result: Dict[str, Any] = {
            "delta_col": delta_col,
            "feature_cols": feature_cols,
            "n_samples": len(data),
            "paths": {},
            "stats": {},
            "summary_df": None,
        }

        # ---------- 4. 整体分析 ----------
        if enable_overall:
            overall_dir = os.path.join(out_dir, "overall") if not compute_stats_only else None
            if overall_dir:
                os.makedirs(overall_dir, exist_ok=True)

            print(f"\n{'=' * 50}")
            print(f"[整体分析] 开始处理全部样本 (n={len(data)})")
            print(f"{'=' * 50}")

            result["stats"]["overall"] = self._run_single_analysis(
                data=data,
                delta_col=delta_col,
                feature_cols=feature_cols,
                time_col=time_col if has_time else None,
                save_dir=overall_dir,
                title_prefix=f"{surface_cn}整体" if surface_cn else "整体",
                plot_univariate=plot_univariate and not compute_stats_only,
                plot_vs_delta=plot_vs_delta and not compute_stats_only,
                plot_time=plot_time and not compute_stats_only,
                sample_for_scatter=sample_for_scatter,
                random_state=random_state,
                figsize_univariate=figsize_univariate,
                compute_stats_only=compute_stats_only,
            )
            if overall_dir:
                result["paths"]["overall"] = overall_dir
            print(f"[整体分析] 完成")
        else:
            print(f"\n[整体分析] 已跳过 (enable_overall=False)")

        # ---------- 5. 分规格分析 ----------
        if enable_by_group and group_col and group_col in data.columns:
            group_dir = os.path.join(out_dir, "by_group") if not compute_stats_only else None
            if group_dir:
                os.makedirs(group_dir, exist_ok=True)

            result["stats"]["by_group"] = {}
            result["paths"]["by_group"] = {}

            print(f"\n{'=' * 50}")
            print(f"[分规格分析] 开始处理分组数据")
            print(f"{'=' * 50}")

            vc = data[group_col].value_counts()
            if target_groups is not None:
                top_groups = [g for g in target_groups if g in vc.index]
                print(f"[分规格] 使用 target_groups 白名单，共 {len(top_groups)} 个组")
            else:
                top_groups = vc.head(max_groups).index.tolist()
                print(f"[分规格] 共 {vc.nunique()} 个规格，分析频次最高的 {len(top_groups)} 个")

            for g in top_groups:
                g_data = data[data[group_col] == g]
                if len(g_data) < 30:
                    print(f"  规格 {g} 样本过少({len(g_data)})，跳过")
                    continue
                g_name = str(g).replace("/", "_").replace("\\", "_")[:40]
                g_save = os.path.join(group_dir, g_name) if group_dir else None
                if g_save:
                    os.makedirs(g_save, exist_ok=True)
                print(f"  -> 规格 {g} (n={len(g_data)})")
                result["stats"]["by_group"][g] = self._run_single_analysis(
                    data=g_data,
                    delta_col=delta_col,
                    feature_cols=feature_cols,
                    time_col=time_col if has_time else None,
                    save_dir=g_save,
                    title_prefix=f"规格[{g}]",
                    plot_univariate=plot_univariate and not compute_stats_only,
                    plot_vs_delta=plot_vs_delta and not compute_stats_only,
                    plot_time=plot_time and not compute_stats_only,
                    sample_for_scatter=sample_for_scatter,
                    random_state=random_state,
                    figsize_univariate=figsize_univariate,
                    compute_stats_only=compute_stats_only,
                )
                if g_save:
                    result["paths"]["by_group"][g] = g_save
            print(f"[分规格分析] 完成")
        elif enable_by_group and not group_col:
            print(f"\n[分规格分析] 已跳过 (未指定 group_col)")
        else:
            print(f"\n[分规格分析] 已跳过 (enable_by_group=False)")

        # ---------- 6. train / test 分布对比 ----------
        if (
            plot_train_test
            and not compute_stats_only
            and train_idx is not None
            and test_idx is not None
        ):
            tt_dir = os.path.join(out_dir, "train_test")
            os.makedirs(tt_dir, exist_ok=True)
            result["stats"]["train_test"] = self._compare_train_test(
                data=data,
                delta_col=delta_col,
                feature_cols=feature_cols,
                train_idx=train_idx,
                test_idx=test_idx,
                save_dir=tt_dir,
                title_prefix=f"{surface_cn}" if surface_cn else "",
            )
            result["paths"]["train_test"] = tt_dir

        # ---------- 7. 建模后残差诊断（兼容旧接口） ----------
        if (
            plot_model_residual
            and not compute_stats_only
            and model_residual_col
            and model_residual_col in work.columns
        ):
            mr_dir = os.path.join(out_dir, "model_residual")
            os.makedirs(mr_dir, exist_ok=True)
            mr_cols = [model_residual_col] + [c for c in feature_cols if c in work.columns]
            if has_time and time_col in work.columns:
                mr_cols.append(time_col)
            mr_data = work[mr_cols].dropna(subset=[model_residual_col])
            print(f"[建模后残差] 有效样本: {len(mr_data)}")
            result["stats"]["model_residual"] = self._run_single_analysis(
                data=mr_data,
                delta_col=model_residual_col,
                feature_cols=[c for c in feature_cols if c in mr_data.columns],
                time_col=time_col if has_time else None,
                save_dir=mr_dir,
                title_prefix=f"{surface_cn}模型残差" if surface_cn else "模型残差",
                plot_univariate=plot_univariate,
                plot_vs_delta=plot_vs_delta,
                plot_time=plot_time,
                sample_for_scatter=sample_for_scatter,
                random_state=random_state,
                figsize_univariate=figsize_univariate,
                compute_stats_only=False,
            )
            result["paths"]["model_residual"] = mr_dir

        # ---------- 8. 统一 summary DataFrame ----------
        result["summary_df"] = self._build_summary_df(result["stats"], delta_col)

        # ========== 导出 Excel 总结 ==========
        if not compute_stats_only and out_dir:
            try:
                excel_path = os.path.join(out_dir, "eda_summary.xlsx")
                self.export_summary_to_excel(
                    result=result,
                    output_path=excel_path,
                    surface=surface_cn if surface_cn else None,
                    group_col=group_col,
                )
            except Exception as e:
                print(f"  [警告] Excel 导出失败: {e}")
        # ===========================================

        if not compute_stats_only:
            print(f"\n[完成] 所有图表与统计已保存至: {out_dir}")
        else:
            print(f"\n[完成] 仅计算统计（未画图）")

        # 打印本次执行摘要
        print(f"\n{'=' * 50}")
        print(f"[执行摘要]")
        print(f"  - 整体分析: {'✅ 已执行' if enable_overall else '⏭️ 已跳过'}")
        print(f"  - 分组分析: {'✅ 已执行' if (enable_by_group and group_col) else '⏭️ 已跳过'}")
        print(f"  - 样本总数: {len(data)}")
        print(f"{'=' * 50}")
        return result

    # ------------------------------------------------------------------
    # 便捷方法：仅对 delta Series 做快速诊断（训练后常用）
    # ------------------------------------------------------------------
    def analyze_residual_series(
        self,
        delta: pd.Series,
        features: Optional[pd.DataFrame] = None,
        time_series: Optional[pd.Series] = None,
        train_idx: Optional[Union[pd.Index, np.ndarray, List]] = None,
        test_idx: Optional[Union[pd.Index, np.ndarray, List]] = None,
        save_dir: Optional[str] = None,
        title_prefix: str = "模型残差",
        sample_for_scatter: Optional[int] = 8000,
        compute_stats_only: bool = False,
        random_state: int = 42,
    ) -> Dict:
        """
        训练后最自然的调用方式：直接传 delta Series + 可选特征矩阵。
        内部构造临时 DataFrame 后复用 analyze。
        """
        delta = delta.dropna()
        tmp = pd.DataFrame({"_delta_": delta})
        feature_cols = []
        if features is not None:
            features = features.reindex(delta.index)
            for c in features.columns:
                tmp[c] = features[c]
            feature_cols = list(features.columns)
        if time_series is not None:
            tmp["Produce Time"] = time_series.reindex(delta.index)

        return self.analyze(
            df=tmp,
            delta_col="_delta_",
            feature_cols=feature_cols if feature_cols else None,
            time_col="Produce Time" if time_series is not None else "Produce Time",
            train_idx=train_idx,
            test_idx=test_idx,
            save_dir=save_dir or os.path.join(self.default_save_dir, "model_residual_series"),
            plot_univariate=True,
            plot_vs_delta=bool(feature_cols),
            plot_time=time_series is not None,
            plot_train_test=train_idx is not None and test_idx is not None,
            plot_model_residual=False,
            compute_stats_only=compute_stats_only,
            sample_for_scatter=sample_for_scatter,
            random_state=random_state,
        )

    # ------------------------------------------------------------------
    # 内部：列名解析（可被 default_feature_map 覆盖）
    # ------------------------------------------------------------------
    def _resolve_columns(
        self,
        df: pd.DataFrame,
        surface: Optional[str],
        lab_col: Optional[str],
        actual_col: Optional[str],
        feature_cols: Optional[Sequence[str]],
        default_feature_map: Optional[Dict[str, List[str]]],
    ):
        surface_cn = ""
        default_features: List[str] = []

        if surface is not None:
            prefix = "Top" if surface == "Top" else "Bot"
            surface_cn = "上" if surface == "Top" else "下"
            if lab_col is None:
                lab_col = f"{surface_cn}表面镀层重量A(XA1_0)"
            if actual_col is None:
                actual_col = f"Tin Weight_Actual[g/m2]_GALV_WEIGHT_{prefix.upper()}_Avg"

            # 优先使用外部传入的 feature map
            if default_feature_map and surface in default_feature_map:
                default_features = list(default_feature_map[surface])
            else:
                default_features = [
                    f"{prefix}_Current_Sum",
                    f"{prefix}_Current_Per_Speed",
                    f"{prefix}_Theoretical_Factor",
                    "Speed[m/min]_Process_Avg",
                    "Dimension_[mm]_Thickness",
                    "Dimension_[mm]_Width",
                    "Steel_Grade_Encoded",
                ]
                if actual_col in df.columns:
                    default_features = [actual_col] + default_features

        if feature_cols is not None:
            default_features = list(feature_cols)

        return lab_col, actual_col, default_features, surface_cn

    # ------------------------------------------------------------------
    # 内部：单次完整分析（整体 or 某个规格）
    # ------------------------------------------------------------------
    def _run_single_analysis(
        self,
        data: pd.DataFrame,
        delta_col: str,
        feature_cols: List[str],
        time_col: Optional[str],
        save_dir: Optional[str],
        title_prefix: str,
        plot_univariate: bool,
        plot_vs_delta: bool,
        plot_time: bool,
        sample_for_scatter: Optional[int],
        random_state: int,
        figsize_univariate: tuple,
        compute_stats_only: bool = False,
    ) -> Dict:
        stats_dict: Dict = {}

        # 1. 测量偏差自身分布（始终算统计；按需画图）
        stats_dict["delta"] = self._describe_series(data[delta_col])
        if plot_univariate and save_dir and not compute_stats_only:
            self._plot_univariate(
                series=data[delta_col],
                name=delta_col,
                save_path=os.path.join(save_dir, "delta_dist.png"),
                title=f"{title_prefix} 测量偏差(Delta)分布",
                figsize=figsize_univariate,
                stats_dict=stats_dict["delta"],
            )

        # 2. 各特征单变量分布
        stats_dict["features"] = {}
        for col in feature_cols:
            if pd.api.types.is_numeric_dtype(data[col]):
                s = self._describe_series(data[col])
                stats_dict["features"][col] = s
                if plot_univariate and save_dir and not compute_stats_only:
                    self._plot_univariate(
                        series=data[col],
                        name=col,
                        save_path=os.path.join(save_dir, f"feature_dist_{self._safe_name(col)}.png"),
                        title=f"{title_prefix} 特征分布: {col}",
                        figsize=figsize_univariate,
                        stats_dict=s,
                    )
            else:
                if plot_univariate and save_dir and not compute_stats_only:
                    self._plot_categorical(
                        series=data[col],
                        name=col,
                        save_path=os.path.join(save_dir, f"feature_dist_{self._safe_name(col)}.png"),
                        title=f"{title_prefix} 特征分布: {col}",
                    )

        # 3. 特征 vs 测量偏差
        if plot_vs_delta and save_dir and not compute_stats_only:
            for col in feature_cols:
                if not pd.api.types.is_numeric_dtype(data[col]):
                    self._plot_residual_by_category(
                        data=data,
                        delta_col=delta_col,
                        cat_col=col,
                        save_path=os.path.join(save_dir, f"delta_vs_{self._safe_name(col)}.png"),
                        title=f"{title_prefix} 测量偏差 vs {col}",
                    )
                else:
                    self._plot_scatter_vs_residual(
                        data=data,
                        delta_col=delta_col,
                        feature_col=col,
                        save_path=os.path.join(save_dir, f"delta_vs_{self._safe_name(col)}.png"),
                        title=f"{title_prefix} 测量偏差 vs {col}",
                        sample_n=sample_for_scatter,
                        random_state=random_state,
                    )

        # 4. 测量偏差 vs 时间
        if plot_time and time_col and time_col in data.columns and save_dir and not compute_stats_only:
            self._plot_delta_vs_time(
                data=data,
                delta_col=delta_col,
                time_col=time_col,
                save_path=os.path.join(save_dir, "delta_vs_time.png"),
                title=f"{title_prefix} 测量偏差随时间变化",
                sample_n=sample_for_scatter,
                random_state=random_state,
            )

         # ========== 提取 LOWESS 摘要（仅当有特征时） ==========
        stats_dict["lowess_summary"] = {}
        for col in feature_cols:
            if pd.api.types.is_numeric_dtype(data[col]):
                stats_dict["lowess_summary"][col] = self._extract_lowess_summary(
                    data=data,
                    delta_col=delta_col,
                    feature_col=col,
                )

        # ========== 提取时间趋势 ==========
        if time_col and time_col in data.columns:
            stats_dict["time_trend"] = self._extract_time_trend(
                data=data,
                delta_col=delta_col,
                time_col=time_col,
            )

        # ========== 提取数据质量 ==========
        all_cols = [delta_col] + feature_cols
        stats_dict["data_quality"] = self._extract_data_quality(
            data=data,
            columns=all_cols,
        )

        return stats_dict

    # ==================================================================
    #  从 LOWESS 拟合提取数值摘要
    # ==================================================================
    def _extract_lowess_summary(
            self,
            data: pd.DataFrame,
            delta_col: str,
            feature_col: str,
            frac: float = 0.3,  # LOWESS 平滑参数，与 seaborn 默认一致
    ) -> Dict:
        """
        从特征与 Delta 的散点中提取 LOWESS 曲线的数值摘要。
        用于判断特征与偏差的关系方向、最优工作点、影响幅度。
        """
        from statsmodels.nonparametric.smoothers_lowess import lowess

        df = data[[feature_col, delta_col]].dropna()
        if len(df) < 10:
            return {
                "trend_direction": "样本不足",
                "extreme_x": np.nan,
                "extreme_delta": np.nan,
                "delta_range": np.nan,
            }

        # 计算 LOWESS
        result = lowess(df[delta_col], df[feature_col], frac=frac, return_sorted=True)
        x_sorted = result[:, 0]
        y_sorted = result[:, 1]

        # 1. 趋势方向判断
        # 用线性回归拟合 LOWESS 曲线，判断整体斜率方向
        slope, intercept = np.polyfit(x_sorted, y_sorted, 1)
        if abs(slope) < 0.001 * (y_sorted.max() - y_sorted.min()) / (x_sorted.max() - x_sorted.min() + 1e-10):
            trend = "平坦"
        elif slope > 0:
            trend = "上升"
        else:
            trend = "下降"

        # 补充：检测 U 型或倒 U 型（二阶导数变化）
        # 用二次拟合判断凹凸性
        if len(x_sorted) > 5:
            coeffs = np.polyfit(x_sorted, y_sorted, 2)
            if coeffs[0] > 0:
                shape = "U型"
            elif coeffs[0] < 0:
                shape = "倒U型"
            else:
                shape = "线性"
            trend = f"{trend}({shape})"

        # 2. 极值点（最小值点，即 Delta 最接近 0 的点）
        abs_min_idx = np.argmin(np.abs(y_sorted))
        extreme_x = x_sorted[abs_min_idx]
        extreme_delta = y_sorted[abs_min_idx]

        # 3. Delta 总变化幅度（特征两端对应的 Delta 差值）
        # 取特征值 5% 分位数和 95% 分位数对应的 Delta 值
        x_low = np.percentile(x_sorted, 5)
        x_high = np.percentile(x_sorted, 95)
        y_at_low = np.interp(x_low, x_sorted, y_sorted)
        y_at_high = np.interp(x_high, x_sorted, y_sorted)
        delta_range = abs(y_at_high - y_at_low)

        return {
            "trend_direction": trend,
            "extreme_x": float(extreme_x),
            "extreme_delta": float(extreme_delta),
            "delta_range": float(delta_range),
        }

    # ==================================================================
    # 从时间序列提取趋势摘要
    # ==================================================================
    def _extract_time_trend(
            self,
            data: pd.DataFrame,
            delta_col: str,
            time_col: str,
    ) -> Dict:
        """
        从 Delta 随时间的变化中提取趋势摘要。
        """
        df = data[[time_col, delta_col]].dropna().sort_values(time_col)
        if len(df) < 10:
            return {
                "trend_slope": np.nan,
                "trend_pvalue": np.nan,
                "trend_significant": False,
                "cpk": np.nan,
                "ppk": np.nan,
            }

        # 1. 趋势斜率 + p-value
        # 将时间转换为数值（天数）
        time_numeric = (df[time_col] - df[time_col].min()).dt.total_seconds() / (24 * 3600)
        slope, intercept, r_value, p_value, std_err = stats.linregress(time_numeric, df[delta_col])

        # 2. CPK / PPK（需要规格限，这里用 3σ 作为近似规格限）
        # 如果数据中没有规格限，使用 ±3σ 作为过程边界
        delta_std = df[delta_col].std()
        delta_mean = df[delta_col].mean()

        # 假设规格限为 3σ（如果没有明确规格限）
        # 注意：实际使用时，如果有明确的规格限（如 ±0.5），应替换
        usl = delta_mean + 3 * delta_std
        lsl = delta_mean - 3 * delta_std

        if delta_std > 0:
            cpu = (usl - delta_mean) / (3 * delta_std)
            cpl = (delta_mean - lsl) / (3 * delta_std)
            cpk = min(cpu, cpl)
            ppk = cpk  # 简化：长期用同一个值
        else:
            cpk = np.nan
            ppk = np.nan

        return {
            "trend_slope": float(slope),
            "trend_pvalue": float(p_value),
            "trend_significant": p_value < 0.05,
            "cpk": float(cpk) if not np.isnan(cpk) else np.nan,
            "ppk": float(ppk) if not np.isnan(ppk) else np.nan,
        }

    # ==================================================================
    # 分组诊断摘要（用于规格组）
    # ==================================================================
    def _extract_group_diagnostics(
            self,
            data: pd.DataFrame,
            delta_col: str,
            group_col: str,
    ) -> pd.DataFrame:
        """
        对每个规格组提取诊断信息：
        - 均值、标准差
        - 均值是否显著 ≠ 0（t 检验）
        """
        rows = []

        for group_name, group_data in data.groupby(group_col):
            if len(group_data) < 10:
                continue

            delta = group_data[delta_col].dropna()
            mean_val = delta.mean()
            std_val = delta.std()

            # t 检验：均值是否显著 ≠ 0
            t_stat, p_val = stats.ttest_1samp(delta, 0)

            rows.append({
                "规格组": str(group_name),
                "样本数": len(delta),
                "Delta均值": float(mean_val),
                "Delta标准差": float(std_val),
                "均值是否显著≠0": p_val < 0.05,
                "t检验p值": float(p_val),
            })

        return pd.DataFrame(rows)

    # ==================================================================
    # 数据质量统计（缺失率 + 异常值占比）
    # ==================================================================
    def _extract_data_quality(
            self,
            data: pd.DataFrame,
            columns: List[str],
    ) -> pd.DataFrame:
        """
        统计指定列的缺失率和异常值占比（IQR 方法）。
        """
        rows = []

        for col in columns:
            if col not in data.columns:
                continue

            series = data[col]
            total = len(series)

            # 缺失率
            missing_rate = series.isna().mean()

            # 异常值占比（仅数值列）
            outlier_rate = np.nan
            if pd.api.types.is_numeric_dtype(series):
                q1 = series.quantile(0.25)
                q3 = series.quantile(0.75)
                iqr = q3 - q1
                lower = q1 - 1.5 * iqr
                upper = q3 + 1.5 * iqr
                if iqr > 0:
                    outliers = ((series < lower) | (series > upper)).sum()
                    outlier_rate = outliers / total

            rows.append({
                "特征名": col,
                "缺失率": float(missing_rate) if not np.isnan(missing_rate) else 0.0,
                "异常值占比": float(outlier_rate) if not np.isnan(outlier_rate) else np.nan,
            })

        return pd.DataFrame(rows)

    # ==================================================================
    # 构建 LOWESS 摘要 DataFrame
    # ==================================================================
    def _build_lowess_excel_df(self, result: Dict, surface: Optional[str] = None) -> pd.DataFrame:
        """构建 特征-偏差关系摘要 DataFrame"""
        rows = []
        surface_label = surface if surface else result.get('surface', 'Unknown')

        # 从整体分析中提取
        overall_stats = result.get('stats', {}).get('overall', {})
        lowess_summary = overall_stats.get('lowess_summary', {})

        for feature_name, summary in lowess_summary.items():
            rows.append({
                '特征名': feature_name,
                '表面': surface_label,
                'LOWESS趋势方向': summary.get('trend_direction', ''),
                '极值点x值': summary.get('extreme_x'),
                '极值点Delta值': summary.get('extreme_delta'),
                'Delta总变化幅度': summary.get('delta_range'),
            })

        return pd.DataFrame(rows)

    # ==================================================================
    # 构建分组诊断 DataFrame
    # ==================================================================
    def _build_group_diagnostic_excel_df(
            self,
            result: Dict,
            group_col: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        构建规格组诊断 DataFrame。
        注意：分组诊断信息在 _run_single_analysis 中并没有按分组保存，
        需要从原始数据中提取，或者在 analyze 方法中传入分组数据。

        这里简化实现：从 result 的 stats 中提取 by_group 的残差统计，
        然后补充 t 检验结果。
        """
        rows = []

        by_group = result.get('stats', {}).get('by_group', {})
        if not by_group:
            return pd.DataFrame()

        for group_name, group_stats in by_group.items():
            residual_stats = group_stats.get('delta', {})
            n = residual_stats.get('count', 0)
            mean_val = residual_stats.get('mean')
            std_val = residual_stats.get('std')

            if n < 10 or mean_val is None or std_val is None:
                continue

            # 计算 t 检验 p-value（用均值和样本量近似）
            # t = mean / (std / sqrt(n))
            if std_val > 0 and n > 1:
                t_stat = mean_val / (std_val / np.sqrt(n))
                # 使用 scipy 计算 p-value
                from scipy import stats as sp_stats
                p_val = sp_stats.t.sf(np.abs(t_stat), n - 1) * 2
                is_significant = p_val < 0.05
            else:
                p_val = np.nan
                is_significant = False

            rows.append({
                '规格组': str(group_name),
                '样本数': n,
                'Delta均值': float(mean_val) if mean_val is not None else np.nan,
                'Delta标准差': float(std_val) if std_val is not None else np.nan,
                '均值是否显著≠0': is_significant,
                't检验p值': float(p_val) if not np.isnan(p_val) else np.nan,
            })

        return pd.DataFrame(rows)

    # ==================================================================
    # 构建时间趋势 DataFrame
    # ==================================================================
    def _build_time_trend_excel_df(self, result: Dict, surface: Optional[str] = None) -> pd.DataFrame:
        """构建时间趋势诊断 DataFrame"""
        rows = []
        surface_label = result.get('surface', 'Unknown')

        # 从整体分析中提取
        overall_stats = result.get('stats', {}).get('overall', {})
        time_trend = overall_stats.get('time_trend', {})

        if not time_trend:
            return pd.DataFrame()

        rows.append({
            '表面': surface_label,
            '趋势斜率': time_trend.get('trend_slope'),
            '趋势p值': time_trend.get('trend_pvalue'),
            '趋势是否显著': time_trend.get('trend_significant', False),
            'CPK': time_trend.get('cpk'),
            'PPK': time_trend.get('ppk'),
        })

        return pd.DataFrame(rows)


    # ------------------------------------------------------------------
    # 统计描述（与绘图解耦）
    # ------------------------------------------------------------------
    @staticmethod
    def _describe_series(series: pd.Series) -> Dict:
        s = series.dropna()
        if len(s) == 0:
            return {
                "count": 0,
                "mean": np.nan,
                "std": np.nan,
                "min": np.nan,
                "q25": np.nan,
                "median": np.nan,
                "q75": np.nan,
                "max": np.nan,
                "skew": np.nan,
                "kurtosis": np.nan,
                "pos_ratio": np.nan,  # 正值比例（Delta > 0 表示在线测量偏低）
                "neg_ratio": np.nan,  # 负值比例（Delta < 0 表示在线测量偏高）
            }
        return {
            "count": int(s.count()),
            "mean": float(s.mean()),
            "std": float(s.std()),
            "min": float(s.min()),
            "q25": float(s.quantile(0.25)),
            "median": float(s.median()),
            "q75": float(s.quantile(0.75)),
            "max": float(s.max()),
            "skew": float(s.skew()),
            "kurtosis": float(s.kurtosis()),
            "pos_ratio": float((s > 0).mean()),  # 正值比例
            "neg_ratio": float((s < 0).mean()),  # 负值比例
        }

    # ------------------------------------------------------------------
    # 统一 summary DataFrame
    # ------------------------------------------------------------------
    def _build_summary_df(self, stats: Dict, delta_col: str) -> pd.DataFrame:
        rows = []

        def _add_row(scope: str, group: Any, residual_stats: Dict):
            if not residual_stats:
                return
            rows.append(
                {
                    "scope": scope,
                    "group": group,
                    "delta_col": delta_col,
                    "n": residual_stats.get("count"),
                    "mean": residual_stats.get("mean"),
                    "std": residual_stats.get("std"),
                    "median": residual_stats.get("median"),
                    "skew": residual_stats.get("skew"),
                    "pos_ratio": residual_stats.get("pos_ratio"),
                    "neg_ratio": residual_stats.get("neg_ratio"),
                    "min": residual_stats.get("min"),
                    "max": residual_stats.get("max"),
                }
            )

        if "overall" in stats and "delta" in stats["overall"]:
            _add_row("overall", None, stats["overall"]["delta"])

        if "by_group" in stats:
            for g, g_stats in stats["by_group"].items():
                if "delta" in g_stats:
                    _add_row("by_group", g, g_stats["delta"])

        if "model_residual" in stats and "delta" in stats["model_residual"]:
            _add_row("model_residual", None, stats["model_residual"]["delta"])

        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)

    # ========== Excel 总结导出方法 ==========
    def export_summary_to_excel(
            self,
            result: Dict[str, Any],
            output_path: str,
            surface: Optional[str] = None,
            group_col: Optional[str] = None,
    ) -> None:
        """
        将 EDA 分析结果导出为格式化的 Excel 文件
        """
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 原有：构建各 Sheet 的数据
        delta_df = self._build_residual_excel_df(result, surface)
        feature_df = self._build_feature_excel_df(result, surface)

        # ====== 新增：构建新增 Sheet 的数据 ======
        lowess_df = self._build_lowess_excel_df(result)
        group_diag_df = self._build_group_diagnostic_excel_df(result, group_col)
        time_trend_df = self._build_time_trend_excel_df(result)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 原有 Sheet
            if not delta_df.empty:
                delta_df.to_excel(writer, sheet_name='偏差量统计汇总', index=False)
            if not feature_df.empty:
                feature_df.to_excel(writer, sheet_name='特征统计汇总', index=False)

            # ====== 新增 Sheet ======
            if not lowess_df.empty:
                lowess_df.to_excel(writer, sheet_name='特征-偏差关系摘要', index=False)
            if not group_diag_df.empty:
                group_diag_df.to_excel(writer, sheet_name='规格组诊断', index=False)
            if not time_trend_df.empty:
                time_trend_df.to_excel(writer, sheet_name='时间趋势诊断', index=False)

        # 应用 Excel 格式
        self._apply_excel_formatting(output_path)
        print(f"  [Excel 导出] 总结表格已保存至: {output_path}")


    def _build_residual_excel_df(self, result: Dict, surface: Optional[str] = None) -> pd.DataFrame:
        """构建偏差量（Delta）统计汇总 DataFrame"""
        rows = []
        surface_label = surface if surface else "Unknown"

        # 1. 整体统计
        overall_stats = result.get('stats', {}).get('overall', {}).get('delta', {})
        if overall_stats:
            rows.append(self._stats_to_row('ALL (整体)', None, overall_stats, surface_label))

        # 2. 分组统计
        by_group = result.get('stats', {}).get('by_group', {})
        for group_name, group_stats in by_group.items():
            residual_stats = group_stats.get('delta', {})
            if residual_stats:
                rows.append(self._stats_to_row(str(group_name), '分组', residual_stats, surface_label))

        # 3. 模型残差（如果有）
        model_residual = result.get('stats', {}).get('model_residual', {}).get('delta', {})
        if model_residual:
            rows.append(self._stats_to_row('模型残差', 'model', model_residual, surface_label))

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        # 列顺序调整
        col_order = ['规格组', 'scope', '表面', '样本数', '均值', '标准差', '中位数',
                     '最小值', '最大值', '偏度', '峰度', '正偏差比例(%)', '负偏差比例(%)']
        existing_cols = [c for c in col_order if c in df.columns]
        return df[existing_cols]

    def _build_feature_excel_df(self, result: Dict, surface: Optional[str] = None) -> pd.DataFrame:
        """构建特征统计汇总 DataFrame"""
        rows = []
        surface_label = surface if surface else "Unknown"

        overall_stats = result.get('stats', {}).get('overall', {}).get('features', {})
        for feature_name, feature_stats in overall_stats.items():
            rows.append({
                '特征名称': feature_name,
                '表面': surface_label,
                '样本数': feature_stats.get('count'),
                '均值': feature_stats.get('mean'),
                '标准差': feature_stats.get('std'),
                '最小值': feature_stats.get('min'),
                '中位数': feature_stats.get('median'),
                '最大值': feature_stats.get('max'),
                '偏度': feature_stats.get('skew'),
                '峰度': feature_stats.get('kurtosis'),
            })

        return pd.DataFrame(rows)

    def _stats_to_row(self, group_name: str, scope: str, stats: Dict, surface: str) -> Dict:
        """将统计字典转换为行数据
        pos_ratio: Delta > 0 的比例（在线测量偏低，需要正向补偿）
        neg_ratio: Delta < 0 的比例（在线测量偏高，需要负向补偿）
        """
        return {
            '规格组': group_name,
            'scope': scope,
            '表面': surface,
            '样本数': stats.get('count'),
            '均值': stats.get('mean'),
            '标准差': stats.get('std'),
            '中位数': stats.get('median'),
            '最小值': stats.get('min'),
            '最大值': stats.get('max'),
            '偏度': stats.get('skew'),
            '峰度': stats.get('kurtosis'),
            '正偏差比例(%)': stats.get('pos_ratio', 0) * 100 if stats.get('pos_ratio') else None,
            '负偏差比例(%)': stats.get('neg_ratio', 0) * 100 if stats.get('neg_ratio') else None,
        }

    def _apply_excel_formatting(self, filepath: str) -> None:
        """应用 Excel 格式（颜色、边框、列宽）"""
        try:
            wb = load_workbook(filepath)

            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row == 1:
                    continue

                # 表头样式
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_alignment

                # 数据边框 + 居中对齐
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                # 自动列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value is not None:
                                cell_len = len(str(cell.value))
                                if cell_len > max_length:
                                    max_length = cell_len
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # 条件格式：正偏差比例（越高越需要关注）
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and '正偏差比例' in str(cell.value):
                        if ws.max_row >= 2:
                            color_scale = ColorScaleRule(
                                start_type='min', start_color='63BE7B',
                                mid_type='percentile', mid_value=50, mid_color='FED976',
                                end_type='max', end_color='F03B20'
                            )
                            ws.conditional_formatting.add(
                                f'{cell.column_letter}2:{cell.column_letter}{ws.max_row}',
                                color_scale
                            )
                        break

            wb.save(filepath)
        except Exception as e:
            print(f"  [警告] Excel 格式应用失败: {e}")

    # ------------------------------------------------------------------
    # 绘图工具
    # ------------------------------------------------------------------
    def _plot_univariate(
            self,
            series: pd.Series,
            name: str,
            save_path: str,
            title: str,
            figsize: tuple = (14, 4),
            stats_dict: Optional[Dict] = None,
    ) -> None:
        s = series.dropna()
        desc = stats_dict if stats_dict is not None else self._describe_series(s)

        fig, axes = plt.subplots(1, 3, figsize=figsize)

        # 1. Histogram + KDE
        sns.histplot(s, kde=True, ax=axes[0], color="steelblue", edgecolor="white")

        mean_val = desc.get("mean", np.nan)
        std_val = desc.get("std", np.nan)
        median_val = desc.get("median", np.nan)

        # 均值线
        if not np.isnan(mean_val):
            axes[0].axvline(mean_val, color="red", ls="--", linewidth=1.5, label=f"mean = {mean_val:.3f}")

        # 标准差（仅做图例展示，用 alpha=0 隐藏线条，实现并列对齐）
        if not np.isnan(std_val):
            axes[0].axvline(mean_val, color="none", label=f"std     = {std_val:.3f}")

        # 中位数线
        if not np.isnan(median_val):
            axes[0].axvline(median_val, color="orange", ls="--", linewidth=1.5, label=f"median = {median_val:.3f}")

        axes[0].set_title("Histogram + KDE")
        axes[0].legend(fontsize=8, loc="upper right")

        # 2. Boxplot
        sns.boxplot(x=s, ax=axes[1], color="lightblue")
        axes[1].set_title("Boxplot")

        # 3. Q-Q Plot
        stats.probplot(s, dist="norm", plot=axes[2])
        axes[2].set_title("Q-Q Plot")

        fig.suptitle(title, fontsize=12)
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {save_path}")


    def _plot_categorical(
        self,
        series: pd.Series,
        name: str,
        save_path: str,
        title: str,
        top_n: int = 20,
    ):
        vc = series.value_counts().head(top_n)
        fig, ax = plt.subplots(figsize=(8, max(3, len(vc) * 0.35)))
        vc.sort_values().plot(kind="barh", ax=ax, color="steelblue")
        ax.set_xlabel("Count")
        ax.set_title(title)
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {save_path}")

    def _plot_scatter_vs_residual(
        self,
        data: pd.DataFrame,
        delta_col: str,
        feature_col: str,
        save_path: str,
        title: str,
        sample_n: Optional[int],
        random_state: int,
    ):
        plot_df = data[[feature_col, delta_col]].dropna()
        if sample_n is not None and len(plot_df) > sample_n:
            plot_df = plot_df.sample(sample_n, random_state=random_state)

        fig, ax = plt.subplots(figsize=(7, 5))
        sns.regplot(
            data=plot_df,
            x=feature_col,
            y=delta_col,
            ax=ax,
            scatter_kws={"alpha": 0.35, "s": 12, "label": "样本点"},
            line_kws={"color": "red", "lw": 2, "label": "LOWESS 拟合曲线"},
            lowess=True,
        )
        ax.axhline(0, color="gray", ls="--", lw=1, label="零线")
        ax.set_title(title)
        ax.legend(loc="best", fontsize=9)
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {save_path}")

    def _plot_residual_by_category(
        self,
        data: pd.DataFrame,
        delta_col: str,
        cat_col: str,
        save_path: str,
        title: str,
        top_n: int = 15,
    ):
        plot_df = data[[cat_col, delta_col]].dropna()
        top_cats = plot_df[cat_col].value_counts().head(top_n).index
        plot_df = plot_df[plot_df[cat_col].isin(top_cats)]

        fig, ax = plt.subplots(figsize=(9, 5))
        order = plot_df.groupby(cat_col)[delta_col].median().sort_values().index
        sns.boxplot(data=plot_df, x=cat_col, y=delta_col, order=order, ax=ax, color="lightblue")
        ax.axhline(0, color="red", ls="--", lw=1)
        ax.set_title(title)
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {save_path}")

    def _plot_delta_vs_time(
        self,
        data: pd.DataFrame,
        delta_col: str,
        time_col: str,
        save_path: str,
        title: str,
        sample_n: Optional[int],
        random_state: int,
    ):
        plot_df = data[[time_col, delta_col]].dropna().sort_values(time_col)
        if sample_n is not None and len(plot_df) > sample_n:
            idx = np.linspace(0, len(plot_df) - 1, sample_n, dtype=int)
            plot_df = plot_df.iloc[idx]

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.scatter(plot_df[time_col], plot_df[delta_col], alpha=0.3, s=8, c="steelblue")
        if len(plot_df) > 50:
            window = max(20, len(plot_df) // 50)
            rolling = plot_df[delta_col].rolling(window, center=True, min_periods=5).mean()
            ax.plot(plot_df[time_col], rolling, color="red", lw=2, label=f"rolling mean (w={window})")
            ax.legend()
        ax.axhline(0, color="gray", ls="--", lw=1)
        ax.set_title(title)
        ax.set_xlabel(time_col)
        ax.set_ylabel(delta_col)
        fig.autofmt_xdate()
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {save_path}")

    # ------------------------------------------------------------------
    # train / test 分布对比
    # ------------------------------------------------------------------
    def _compare_train_test(
        self,
        data: pd.DataFrame,
        delta_col: str,
        feature_cols: List[str],
        train_idx,
        test_idx,
        save_dir: str,
        title_prefix: str,
    ) -> Dict:
        train_data = data.loc[data.index.intersection(train_idx)]
        test_data = data.loc[data.index.intersection(test_idx)]
        print(f"[Train/Test] train={len(train_data)}, test={len(test_data)}")

        stats_out = {"train_n": len(train_data), "test_n": len(test_data), "ks": {}}

        # 残差分布对比
        fig, ax = plt.subplots(figsize=(8, 4))
        sns.kdeplot(train_data[delta_col].dropna(), ax=ax, label="Train", fill=True, alpha=0.4)
        sns.kdeplot(test_data[delta_col].dropna(), ax=ax, label="Test", fill=True, alpha=0.4)
        ax.set_title(f"{title_prefix} delta Train vs Test")
        ax.legend()
        plt.tight_layout()
        path = os.path.join(save_dir, "residual_train_vs_test.png")
        fig.savefig(path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        print(f"  [保存] {path}")

        # KS 检验 + 特征分布叠加
        for col in [delta_col] + feature_cols:
            if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
                continue
            a = train_data[col].dropna()
            b = test_data[col].dropna()
            if len(a) < 10 or len(b) < 10:
                continue
            try:
                ks_stat, ks_p = stats.ks_2samp(a, b)
            except Exception:
                continue
            stats_out["ks"][col] = {"statistic": float(ks_stat), "pvalue": float(ks_p)}

            fig, ax = plt.subplots(figsize=(7, 4))
            sns.kdeplot(a, ax=ax, label="Train", fill=True, alpha=0.4)
            sns.kdeplot(b, ax=ax, label="Test", fill=True, alpha=0.4)
            ax.set_title(f"{title_prefix} {col}\nKS={ks_stat:.3f}, p={ks_p:.3g}")
            ax.legend()
            plt.tight_layout()
            path = os.path.join(save_dir, f"dist_{self._safe_name(col)}_train_vs_test.png")
            fig.savefig(path, dpi=200, bbox_inches="tight")
            plt.close(fig)

        return stats_out

    # ------------------------------------------------------------------
    # 工具
    # ------------------------------------------------------------------
    @staticmethod
    def _safe_name(name: str) -> str:
        return (
            str(name)
            .replace("/", "_")
            .replace("\\", "_")
            .replace(" ", "_")
            .replace("[", "")
            .replace("]", "")
            .replace("(", "")
            .replace(")", "")
            [:60]
        )




def get_feature_cols(surface: str) -> List[str]:
    """获取默认特征列（与主流程保持一致）"""
    prefix = "Top" if surface == "Top" else "Bot"
    return [
        f"Tin Weight_Actual[g/m2]_GALV_WEIGHT_{prefix.upper()}_Avg",
        f'{prefix}_Weight_Deviation',
        f"{prefix}_Current_Sum",
        f"{prefix}_Current_Per_Speed",
        f"{prefix}_Theoretical_Factor",
        "Speed[m/min]_Process_Avg",
        "Dimension_[mm]_Width",
        "Dimension_[mm]_Thickness",
        "Steel_Grade_Encoded",
    ]

# ==========================================
# 独立的全量 EDA 分析入口（供主流程调用）
# ==========================================
def run_global_eda(
        df: pd.DataFrame,
        surface_list: Optional[List[str]] = None,
        group_col: str = "Setpoint_Group_Label",
        target_groups: Optional[List[str]] = None,
        min_samples: int = 200,
        save_root: str = "result/eda_analysis",
        time_col: str = "Produce Time",
        max_groups: int = 20,
        feature_cols_override: Optional[Dict[str, List[str]]] = None,
        enable_overall: bool = True,
        enable_by_group: bool = True,
        plot_train_test: bool = False,
        plot_model_residual: bool = False,
) -> Dict[str, Any]:
    """
    对全量数据做训练前的 EDA 分析（支持多表面、分规格组）
    """
    if surface_list is None:
        surface_list = ["Top", "Bot"]

    results = {}

    for surface in surface_list:
        print(f"\n{'=' * 60}")
        print(f"[全量 EDA] 开始处理表面: {surface}")
        print(f"{'=' * 60}")

        # 1. 拷贝数据并衍生特征
        work = df.copy()
        prefix = "Top" if surface == "Top" else "Bot"
        speed_col = "Speed[m/min]_Process_Avg"
        current_col = f"{prefix}_Current_Sum"
        per_speed = f"{prefix}_Current_Per_Speed"
        weight_deviation = f"{prefix}_Weight_Deviation"
        online_col = f'Tin Weight_Actual[g/m2]_GALV_WEIGHT_{prefix.upper()}_Avg'
        setpoint_weight = f'Tin Weight_Setpoints[g/m2]_GALV_WEIGHT_{prefix.upper()}_Avg'

        # 衍生 Current_Per_Speed
        if current_col in work.columns and speed_col in work.columns:
            if per_speed not in work.columns:
                work[per_speed] = work[current_col] / (work[speed_col] + 1e-5)

        # 衍生 Weight_Deviation
        if online_col in work.columns and setpoint_weight in work.columns:
            if weight_deviation not in work.columns:
                work[weight_deviation] = work[setpoint_weight] - work[online_col]

        # 2. 确定特征列
        if feature_cols_override and surface in feature_cols_override:
            feature_cols = feature_cols_override[surface]
        else:
            feature_cols = get_feature_cols(surface)

        # 3. 确定 target_groups（自动筛选达标组）
        if target_groups is None:
            if group_col in work.columns:
                group_sizes = work.groupby(group_col).size()
                target_groups = group_sizes[group_sizes >= min_samples].index.tolist()
                print(f"[全量 EDA] 自动筛选达标组: {len(target_groups)} 个")
            else:
                target_groups = None

        # 4. 创建 EDA 分析器并运行
        save_dir = os.path.join(save_root, surface)
        eda = SurfaceEDAAnalyzer(default_save_dir=save_dir)

        result = eda.analyze(
            df=work,
            surface=surface,
            feature_cols=feature_cols,
            group_col=group_col if enable_by_group else None,
            target_groups=target_groups,
            time_col=time_col,
            plot_train_test=plot_train_test,
            plot_model_residual=plot_model_residual,
            max_groups=max_groups,
            enable_overall=enable_overall,
            enable_by_group=enable_by_group,
        )

        results[surface] = result
        print(f"[全量 EDA] {surface} 完成，结果保存至: {save_dir}")

    # ================================================================
    # 在所有 surface 分析完成后，生成综合 Excel 报告（增强版）
    # ================================================================
    try:
        from openpyxl import Workbook
        combined_path = os.path.join(save_root, "eda_summary_combined.xlsx")

        with pd.ExcelWriter(combined_path, engine='openpyxl') as writer:
            for surface, result in results.items():
                # ---- 原有 Sheet ----
                # 1. 偏差量统计汇总
                delta_df = eda._build_residual_excel_df(result, surface)
                if not delta_df.empty:
                    delta_df.to_excel(writer, sheet_name=f'{surface}_偏差量统计', index=False)

                # 2. 特征统计汇总
                feature_df = eda._build_feature_excel_df(result, surface)
                if not feature_df.empty:
                    feature_df.to_excel(writer, sheet_name=f'{surface}_特征统计', index=False)


                # ---- 新增 Sheet ----
                # 4. 特征-偏差关系摘要（LOWESS）
                lowess_df = eda._build_lowess_excel_df(result, surface)
                if not lowess_df.empty:
                    lowess_df.to_excel(writer, sheet_name=f'{surface}_特征-偏差关系', index=False)

                # 5. 规格组诊断（仅当有分组数据时）
                if enable_by_group and group_col:
                    group_diag_df = eda._build_group_diagnostic_excel_df(result, group_col)
                    if not group_diag_df.empty:
                        group_diag_df.to_excel(writer, sheet_name=f'{surface}_规格组诊断', index=False)

                # 6. 时间趋势诊断（仅当有时间列且有整体分析时）
                if enable_overall and time_col:
                    time_trend_df = eda._build_time_trend_excel_df(result)
                    if not time_trend_df.empty:
                        time_trend_df.to_excel(writer, sheet_name=f'{surface}_时间趋势', index=False)

        print(f"\n[综合 Excel] 已保存至: {combined_path}")

    except Exception as e:
        import traceback
        print(f"[警告] 综合 Excel 导出失败: {e}")
        traceback.print_exc()

    return results


# ----------------------------------------------------------------------
# 简单使用示例
# ----------------------------------------------------------------------
if __name__ == "__main__":
    if __name__ == "__main__":
        print("SurfaceEDAAnalyzer 已加载。")
        # 测试代码(对全量数据做 EDA 分析)
        featured_df=pd.read_excel("result/data/feature_engineered_data/featured_data.xlsx")
        run_global_eda(df=featured_df,save_root="result/eda_analysis")
