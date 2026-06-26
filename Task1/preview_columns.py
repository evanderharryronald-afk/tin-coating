"""
preview_columns.py
------------------
预览主表所有列的"列字母 -> 生成列名"对照表。
用途：填写 config.yaml 之前，先跑这个脚本确认要保留哪些列。

用法：
    python preview_columns.py                      # 输出到终端
    python preview_columns.py --out cols.txt       # 同时保存到文件
    python preview_columns.py --filter "Tin"       # 只显示含关键词的行（不区分大小写）
"""

import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── 配置（与 run_merge.py 保持一致，无需改动）──────────────────────────────
HEADER_ROW_START = 11   # L1 所在行（1-indexed）
HEADER_ROW_END   = 14   # L4 所在行（1-indexed）
DATA_START_ROW   = 15   # 数据起始行
# ─────────────────────────────────────────────────────────────────────────────


def build_merge_map(ws):
    """把所有合并单元格范围展开成 {(row,col): 左上角值} 的字典。"""
    m = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                m[(r, c)] = top_val
    return m


def get_val(ws, merge_map, row, col):
    return merge_map.get((row, col), ws.cell(row=row, column=col).value)


def ffill(lst):
    """向右前向填充（跳过 None 和空串）。"""
    result, last = [], None
    for v in lst:
        sv = str(v).strip() if v is not None else ''
        if sv:
            last = sv
        result.append(last)
    return result


def clean(s):
    """清理列名中的换行符和多余空格。"""
    if s is None:
        return ''
    return str(s).replace('\n', '').replace('\r', '').strip()


def detect_max_col(ws, merge_map, data_start_row=15, sample_rows=5):
    """通过扫描数据行确定实际有数据的最大列数。"""
    max_col = 0
    count = 0
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        for c, v in enumerate(row, 1):
            if v is not None:
                max_col = max(max_col, c)
        count += 1
        if count >= sample_rows:
            break
    # 再看表头行
    for r in range(HEADER_ROW_START, HEADER_ROW_END + 1):
        for c in range(1, 2000):
            if get_val(ws, merge_map, r, c) is not None:
                max_col = max(max_col, c)
    return max_col


def generate_column_names(ws, merge_map, max_col):
    """生成所有列的最终列名（含去重后缀）。"""
    # 读取 4 层表头，L4 用合并展开，L1-L3 额外 ffill
    raw = []
    for row_idx in range(HEADER_ROW_START, HEADER_ROW_END + 1):
        raw.append([get_val(ws, merge_map, row_idx, c) for c in range(1, max_col + 1)])

    l1f = ffill(raw[0])
    l2f = ffill(raw[1])
    l3f = ffill(raw[2])
    l4  = [clean(v) for v in raw[3]]   # L4 已由合并展开，不 ffill

    col_names = []
    for i in range(max_col):
        parts = []
        for lvl in [l1f[i], l2f[i], l3f[i], l4[i]]:
            sv = clean(lvl)
            if sv and (not parts or parts[-1] != sv):
                parts.append(sv)
        name = '_'.join(parts) if parts else f'col_{i+1:03d}'
        col_names.append(name)

    # 去重加后缀
    from collections import Counter
    counts = Counter()
    final = []
    for name in col_names:
        counts[name] += 1
        final.append(f"{name}_{counts[name]-1}" if counts[name] > 1 else name)
    return final


def main():
    parser = argparse.ArgumentParser(description="预览主表列字母→列名对照")
    parser.add_argument("xlsx", nargs='?', default=None, help="主表路径（默认从 config.yaml 读取）")
    parser.add_argument("--out", default=None, help="同时输出到文件")
    parser.add_argument("--filter", default=None, help="只显示包含此关键词的行（不区分大小写）")
    args = parser.parse_args()

    # 确定主表路径
    xlsx_path = args.xlsx
    if xlsx_path is None:
        try:
            import yaml
            cfg_path = Path(__file__).parent / "config.yaml"
            with open(cfg_path, encoding='utf-8') as f:
                cfg = yaml.safe_load(f)
            xlsx_path = cfg['master']['path']
        except Exception:
            print("❌ 未指定文件路径，且无法从 config.yaml 读取。")
            print("   用法: python preview_columns.py your_file.xlsx")
            sys.exit(1)

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"❌ 文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"📂 正在读取: {xlsx_path.name}  (合并单元格展开中，稍候...)")
    wb = load_workbook(str(xlsx_path), read_only=False, data_only=True)
    ws = wb.active
    merge_map = build_merge_map(ws)

    max_col = detect_max_col(ws, merge_map)
    col_names = generate_column_names(ws, merge_map, max_col)

    lines = []
    header = f"{'列字母':<6}  {'列号':>5}  列名"
    sep    = "-" * 80
    lines.append(header)
    lines.append(sep)

    kw = args.filter.lower() if args.filter else None
    for i, name in enumerate(col_names):
        letter = get_column_letter(i + 1)
        col_no = i + 1
        line = f"{letter:<6}  {col_no:>5}  {name}"
        if kw and kw not in name.lower():
            continue
        lines.append(line)

    lines.append(sep)
    lines.append(f"共 {max_col} 列")

    output = '\n'.join(lines)
    print(output)

    if args.out:
        out_path = Path(args.out)
        out_path.write_text(output, encoding='utf-8')
        print(f"\n✅ 已保存到: {out_path}")


if __name__ == '__main__':
    main()
