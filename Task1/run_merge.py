"""
run_merge.py (高性能重构版)
------------
读取 config.yaml，执行主副表合并，输出结果 xlsx。
优化策略：前置剪枝过滤、Pandas向量化聚合、内存效率调优。
"""
import pandas as pd
import sys
import argparse
import warnings
from pathlib import Path
from collections import Counter
import time
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import yaml
import json
import hashlib

# ══════════════════════════════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════════════════════════════

def log(msg):
    print(msg, flush=True)


def warn(msg):
    print(f"⚠️  WARNING: {msg}", flush=True)


def die(msg):
    print(f"❌ ERROR: {msg}", flush=True)
    sys.exit(1)


def build_merge_map(ws):
    """展开所有合并单元格：{(row, col): 左上角值}。"""
    m = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                m[(r, c)] = top_val
    return m


def get_val(ws, merge_map, row, col):
    return merge_map.get((row, col), ws.cell(row=row, column=col).value)


def ffill(lst):
    result, last = [], None
    for v in lst:
        sv = str(v).strip() if v is not None else ''
        if sv:
            last = sv
        result.append(last)
    return result


def clean_str(s):
    if s is None:
        return ''
    return str(s).replace('\n', '').replace('\r', '').strip()


def detect_max_col(ws, merge_map):
    """从合并单元格map和表头行直接取最大列号。"""
    max_col = max((c for r, c in merge_map.keys()
                   if HEADER_ROW_START <= r <= HEADER_ROW_END), default=0)
    for r in range(HEADER_ROW_START, HEADER_ROW_END + 1):
        for c in range(max_col, max_col + 20):
            v = ws.cell(row=r, column=c).value
            if v is not None and clean_str(v):
                max_col = max(max_col, c)
    return max_col


def generate_column_names(ws, merge_map, max_col):
    """生成主表所有列的最终列名。"""
    raw = []
    for row_idx in range(HEADER_ROW_START, HEADER_ROW_END + 1):
        raw.append([get_val(ws, merge_map, row_idx, c) for c in range(1, max_col + 1)])

    l1f = ffill(raw[0])
    l2f = ffill(raw[1])
    l3f = ffill(raw[2])
    l4 = [clean_str(v) for v in raw[3]]

    col_names = []
    for i in range(max_col):
        parts = []
        for lvl in [l1f[i], l2f[i], l3f[i], l4[i]]:
            sv = clean_str(lvl)
            if sv and (not parts or parts[-1] != sv):
                parts.append(sv)
        col_names.append('_'.join(parts) if parts else f'col_{i + 1:03d}')

    counts = Counter()
    final = []
    for name in col_names:
        counts[name] += 1
        final.append(f"{name}_{counts[name] - 1}" if counts[name] > 1 else name)
    return final


# ══════════════════════════════════════════════════════════════════════════════
#  主表读取与缓存优化
# ══════════════════════════════════════════════════════════════════════════════

def parse_master_schema(path: Path, keep_col_letters: list[str], base_dir: Path):
    """展开合并单元格、生成列名、应用独立配置文件更正，最后确定 keep_indices。"""
    log(f"📐 解析主表结构: {path.parent.name}/{path.name}")
    log("   展开合并单元格中...")
    wb_full = load_workbook(str(path), read_only=False, data_only=True)
    ws_full = wb_full.active
    merge_map = build_merge_map(ws_full)
    max_col = detect_max_col(ws_full, merge_map)
    log(f"   检测到 {max_col} 列")
    all_col_names = generate_column_names(ws_full, merge_map, max_col)
    wb_full.close()

    patch_path = base_dir / "rename_patch.yaml"
    if patch_path.exists():
        try:
            with open(patch_path, encoding='utf-8') as f:
                rename_patch = yaml.safe_load(f)
            if rename_patch and isinstance(rename_patch, dict):
                for idx, name in enumerate(all_col_names):
                    if name in rename_patch:
                        correct_name = rename_patch[name]
                        log(f"   🔧 检测到独立更正配置：[{name}] -> [{correct_name}]")
                        all_col_names[idx] = correct_name
        except Exception as e:
            warn(f"加载列名更正文件 rename_patch.yaml 失败: {e}")

    keep_indices = []
    col_letter_map = {}
    for letter in keep_col_letters:
        letter = letter.strip().upper()
        idx = column_index_from_string(letter) - 1
        if idx >= max_col:
            die(f"列 {letter} 超出主表范围（共 {max_col} 列）")
        semantic_name = all_col_names[idx]
        keep_indices.append(idx)
        col_letter_map[semantic_name] = letter

    selected_names = [all_col_names[i] for i in keep_indices]
    log(f"   ✅ 结构解析完成，保留 {len(keep_indices)} 列")
    return keep_indices, selected_names, col_letter_map


def get_master_schema_with_cache(master_path: Path, keep_col_letters: list[str], base_dir: Path):
    """带本地文件缓存的表头结构获取函数。"""
    cache_path = base_dir / "master_schema_cache.json"
    patch_path = base_dir / "rename_patch.yaml"

    patch_content = ""
    if patch_path.exists():
        patch_content = patch_path.read_text(encoding='utf-8')

    config_str = f"{master_path.name}_{HEADER_ROW_START}_{HEADER_ROW_END}_{','.join(keep_col_letters)}_{patch_content}"
    config_hash = hashlib.md5(config_str.encode('utf-8')).hexdigest()

    if cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            if cache_data.get("config_hash") == config_hash:
                log("🚀 发现相匹配的主表结构缓存，秒级加载成功（已包含更正配置）")
                return cache_data["keep_indices"], cache_data["selected_names"], cache_data["col_letter_map"]
        except Exception:
            warn("读取结构缓存失败，将重新解析主表。")

    keep_indices, selected_names, col_letter_map = parse_master_schema(master_path, keep_col_letters, base_dir)

    try:
        cache_data = {
            "config_hash": config_hash,
            "keep_indices": keep_indices,
            "selected_names": selected_names,
            "col_letter_map": col_letter_map
        }
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        log("💾 修正后的新结构已重新成功缓存到本地。")
    except Exception as e:
        warn(f"创建结构缓存失败: {e}")

    return keep_indices, selected_names, col_letter_map


def load_master_data(path: Path, keep_indices: list, selected_names: list, col_letter_map: dict) -> pd.DataFrame:
    """流式只读数据行。"""
    log(f"📂 读取主表数据: {path.parent.name}/{path.name}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(v is None for v in row):
            break
        rows.append([row[i] if i < len(row) else None for i in keep_indices])

    wb.close()

    if not rows:
        die(f"主表 {path.parent.name}/{path.name} 没有读取到任何数据行。")

    df = pd.DataFrame(rows, columns=selected_names)
    df.attrs['col_letter_map'] = col_letter_map
    log(f"   ✅ 读取完成：{len(df)} 行 × {len(df.columns)} 列")
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  副表读取 & 高性能合并
# ══════════════════════════════════════════════════════════════════════════════

def load_sub(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in ('.xlsx', '.xls'):
        return pd.read_excel(str(path), header=0)
    elif suffix == '.csv':
        try:
            return pd.read_csv(str(path), header=0, encoding='utf-8')
        except UnicodeDecodeError:
            return pd.read_csv(str(path), header=0, encoding='gbk')
    else:
        die(f"副表格式不支持: {path.suffix}，仅支持 xlsx/xls/csv。")


def merge_one_sub(master_df: pd.DataFrame, sub_cfg: dict, base_dir: Path) -> pd.DataFrame:
    """
    高性能副表合并核心：前置剪枝过滤 + 向量化矩阵聚合。
    """
    sub_path = base_dir / sub_cfg['path']
    if not sub_path.exists():
        die(f"副表文件不存在: {sub_path}")

    log(f"\n📂 读取副表: {sub_path.name}")
    sub_df = load_sub(sub_path)
    log(f"   副表原始: {len(sub_df)} 行 × {len(sub_df.columns)} 列")

    # ── 1. 关键字段定位 ──
    join_key = sub_cfg['join_key']
    master_key_letter = join_key['master_col'].strip().upper()
    sub_key_col = join_key['sub_col'].strip()

    master_key_col = None
    for col in master_df.columns:
        if master_df.attrs.get('col_letter_map', {}).get(col) == master_key_letter:
            master_key_col = col
            break
    if master_key_col is None:
        die(f"主表中找不到列字母 '{master_key_letter}' 对应的列，请确认在 keep_columns 中。")

    if sub_key_col not in sub_df.columns:
        die(f"副表 '{sub_path.name}' 中找不到关键字段列 '{sub_key_col}'")

    # 强制将两表的关键字段转换为字符串类型，避免因隐式类型不一致导致无法匹配
    master_df[master_key_col] = master_df[master_key_col].astype(str).str.strip()
    sub_df[sub_key_col] = sub_df[sub_key_col].astype(str).str.strip()

    # ── 2. ⚡ 核心优化：前置剪枝过滤（降低 groupby 规模） ──
    required_cols_cfg = join_key.get('required_columns', [])

    # 规则 A: 如果配置了特定列必填，先在副表中删掉这些列为空的行
    if required_cols_cfg:
        valid_req_cols = [c for c in required_cols_cfg if c in sub_df.columns]
        if valid_req_cols:
            before_drop = len(sub_df)
            sub_df = sub_df.dropna(subset=valid_req_cols, how='any')
            log(f"   ⚡ [前置剪枝] 根据必填列过滤：副表从 {before_drop} 行减少至 {len(sub_df)} 行")

    # 规则 B: 过滤掉不在主表数据集中的无用钢卷号
    master_keys_set = set(master_df[master_key_col].unique())
    before_drop_keys = len(sub_df)
    sub_df = sub_df[sub_df[sub_key_col].isin(master_keys_set)]
    log(f"   ⚡ [前置剪枝] 过滤未匹配主表的钢卷：副表进一步收敛至 {len(sub_df)} 行")

    if len(sub_df) == 0:
        warn("副表过滤后没有剩余有效行，直接返回主表。")
        return master_df

    # ── 3. 策略映射与自动映射字典 ──
    keep_cols_cfg = sub_cfg.get('keep_columns', [])
    if not keep_cols_cfg:
        warn(f"副表 '{sub_path.name}' 未配置 keep_columns。")
        return master_df

    # 将原有配置提取为 Pandas 向量化识别的聚合映射字典
    agg_dict = {}
    rename_dict = {}
    master_cols = set(master_df.columns)

    # 策略合法性静态映射映射（去除低效的原生自定义函数逻辑）
    strategy_map = {
        'mean': 'mean',
        'sum': 'sum',
        'max': 'max',
        'min': 'min',
        'first': 'first',
        'first_nonempty': 'first'  # 由于前面已经进行了 dropna 剪枝，first_nonempty 等价于 first
    }

    for item in keep_cols_cfg:
        col_name = item['col'].strip()
        if col_name not in sub_df.columns:
            die(f"副表中找不到列 '{col_name}'")

        strategy = item.get('strategy', 'mean')
        agg_op = strategy_map.get(strategy, 'mean')

        # 特殊处理文本去重拼接
        if strategy == 'join_unique':
            agg_dict[col_name] = lambda x: ','.join(x.dropna().astype(str).unique())
        else:
            agg_dict[col_name] = agg_op

        if col_name in master_cols and col_name != master_key_col:
            rename_dict[col_name] = f"{col_name}_sub"

    # ── 4. ⚡ 核心优化：向量化矩阵聚合（替代 Python 原生 for 循环） ──
    log("   ⚡ 正在执行底层向量化分组聚合计算...")
    agg_df = sub_df.groupby(sub_key_col, as_index=False).agg(agg_dict)

    if rename_dict:
        agg_df = agg_df.rename(columns=rename_dict)

    # ── 5. 执行合并 (JOIN) ──
    unmatched_action = join_key.get('unmatched', 'keep').strip().lower()
    join_how = 'inner' if unmatched_action == 'drop' else 'left'

    result = master_df.merge(agg_df, left_on=master_key_col, right_on=sub_key_col, how=join_how)

    # 移除多余的关联键
    if sub_key_col != master_key_col and sub_key_col in result.columns:
        result = result.drop(columns=[sub_key_col])

    result.attrs['col_letter_map'] = master_df.attrs.get('col_letter_map', {})
    log(f"   ✅ 合并完成：最终保留 {len(result)} 行")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  主流程
# ══════════════════════════════════════════════════════════════════════════════

def main():
    global HEADER_ROW_START, HEADER_ROW_END, DATA_START_ROW
    t_start = time.time()
    parser = argparse.ArgumentParser(description="高性能主副表合并工具")
    parser.add_argument('--config', default=None, help="配置文件路径")
    args = parser.parse_args()

    if args.config:
        cfg_path = Path(args.config)
    else:
        cfg_path = Path(__file__).parent / "config.yaml"

    if not cfg_path.exists():
        die(f"配置文件不存在: {cfg_path}")

    with open(cfg_path, encoding='utf-8') as f:
        cfg = yaml.safe_load(f)

    base_dir = cfg_path.parent

    # ── 读取主表 ──
    master_cfg = cfg['master']
    search_root = Path(master_cfg['search_root'])
    pattern = master_cfg.get('search_pattern', f"*/{master_cfg['filename']}")
    master_paths = sorted(search_root.glob(pattern))
    if not master_paths:
        die(f"在 {search_root} 下按 '{pattern}' 未找到任何主表文件。")

    log(f"共找到 {len(master_paths)} 个主表文件：")

    HEADER_ROW_START = master_cfg.get('header_row_start', 11)
    HEADER_ROW_END = master_cfg.get('header_row_end', 14)
    DATA_START_ROW = master_cfg.get('data_start_row', 15)

    keep_col_letters = cfg['master'].get('keep_columns', [])
    if not keep_col_letters:
        die("config.yaml 中 master.keep_columns 为空，请至少指定一列。")

    sub_tables = cfg.get('sub_tables', [])

    # 获取表头结构（带本地文件缓存+更正字典联动检测）
    keep_indices, selected_names, col_letter_map = get_master_schema_with_cache(
        master_paths[0], keep_col_letters, base_dir
    )

    all_results = []
    for master_path in master_paths:
        result_df = load_master_data(master_path, keep_indices, selected_names, col_letter_map)
        all_results.append(result_df)

    log(f"纵向拼接 {len(all_results)} 张表...")
    final_df = pd.concat(all_results, ignore_index=True)
    final_df.attrs['col_letter_map'] = col_letter_map

    for i, sub_cfg in enumerate(sub_tables, 1):
        log(f"\n── 合并第 {i} 个副表 ──")
        final_df = merge_one_sub(final_df, sub_cfg, base_dir)

    # ── 写出结果 ──
    out_cfg = cfg.get('output', {})
    out_path = base_dir / out_cfg.get('path', 'merged_result.xlsx')
    out_sheet = out_cfg.get('sheet_name', 'Result')

    log(f"\n💾 写出结果到: {out_path.name}")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # 采用标准高速 openpyxl 引擎写出
    with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=out_sheet, index=False)

    log(f"\n✅ 全部完成！输出文件: {out_path}")
    log(f"   结果表: {len(final_df)} 行 × {len(final_df.columns)} 列")
    log(f"   总耗时: {time.time() - t_start:.1f} 秒")


if __name__ == '__main__':
    main()